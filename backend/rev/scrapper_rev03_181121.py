
#! Importações
from bs4 import BeautifulSoup as bs
from openpyxl.styles import Border, Side, Alignment 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
import openpyxl as op
from openpyxl.drawing.image import Image
import openpyxl.styles as ops
import requests 

class Scrapper():
    def __init__(self):
        self.login = 'camaragibe@projete5d.com.br'
        self.password = '@timeprojete5d'
        self.longinurl = ('https://join.bimcollab.com/WebApp/Account/Login.aspx')
        self.planilha = 'registros/teste2.xlsx'
        Scrapper.coleta_de_dados(self)
        Scrapper.salvar_dados(self)
    def coleta_de_dados(self):
        #*Navegador
        conf = Options()
        conf.add_argument('--headless')
        navegador = webdriver.Chrome(options=conf)
        navegador.get(self.longinurl)

        #* Movimentação no sistema
        navegador.find_element_by_id('UserName').send_keys(self.login)
        navegador.find_element_by_id('Password').send_keys(self.password)
        navegador.find_element_by_css_selector('#LoginButton').click()
        sleep(2)
        selecionar_projeto = navegador.find_element_by_css_selector('#ProjectImage').click()
        selecionar_issue = navegador.find_element_by_css_selector('#HyperLinkIssue').click()

        #*Variáveis
        c=0
        self.descricao = []
        self.titulos = []
        self.prioridades = []
        d = ''

        #*Coleta das informações
        self.lista_issue = navegador.find_elements_by_css_selector(".gridViewRow > .colTitle")
        for i in range(len(self.lista_issue)):
            navegador.find_element_by_css_selector(f"#LabelIndex_{i}").click()
            site = bs(navegador.page_source, 'html.parser')

            #* Coleta do título e prioridade
            titulo = site.find('span',attrs={'id':'LabelDescription'})
            if (titulo.text == "-"):
                navegador.back()
                self.verificacao = True
                continue

            nome = f'registros/{c+1}.jpg'

            prioridade = site.find('span', attrs={'id':'LabelPriority'})
            self.titulos.append(titulo.text)
            self.prioridades.append(prioridade.text)

            #* Coleta da descrição
            u = site.find_all('div',attrs={'class':'onerow commentContent innerPaddingComment'})
            for i in range(len(u)):
                desc = site.find_all('div',attrs={'class':'onerow commentContent innerPaddingComment'})[i]
                d = desc.p.text +'\n'+ d 

            self.descricao.append(d)    
            d = ''

            #* Coleta da imagem
            imagem = site.find('div',attrs={'class':'thumbnailImageFrame backgroundWhite'})
            a =imagem['style'] 
            f = open(nome,'wb')
            response = requests.get('https://join.bimcollab.com/' + a[21:-5])
            f.write(response.content)
            f.close
        
            navegador.back()
            sleep(0.5)
            
            
            c += 1
            if c == 5:
                break
    def salvar_dados(self):
        lw = op.Workbook()
        sheet = lw.active

        k = 1
        r = 2
        t = 3
        x = 8
        n = 10
        for i in range(len(self.titulos)):

            if self.verificacao:
                self.verificacao = False
                continue

            thick = Side(border_style='medium',color='00000000')
            fonte = ops.Font(name='Montserrat',size=11)
            alinhamento = Alignment(horizontal="distributed",vertical="distributed")
            alinhamento2 = Alignment(horizontal="distributed",vertical="distributed",text_rotation=90)

            #?C2:I2
            sheet[f'C{r}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)            
            sheet[f'C{r}'].font = fonte
            sheet[f'C{r}'].alignment = alinhamento
            #? A1
            sheet[f'A{k}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{k}'].font = fonte
            sheet[f'A{k}'].alignment = alinhamento
            #? A2:B2
            sheet[f'A{r}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{r}'].font = fonte
            sheet[f'A{r}'].alignment = alinhamento
            #? B3:K8
            sheet[f'B{t}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{t}'].font = fonte
            sheet[f'B{t}'].alignment = alinhamento

            sheet[f'B{k}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{k}'].font = fonte
            sheet[f'B{k}'].alignment = alinhamento

            sheet[f'J{k}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'J{k}'].font = fonte
            sheet[f'J{k}'].alignment = alinhamento

            sheet[f'J{r}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'J{r}'].font = fonte
            sheet[f'J{r}'].alignment = alinhamento

            sheet[f'A{t}'].border = Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{t}'].font = fonte
            sheet[f'A{t}'].alignment = alinhamento2

            #* Estilo de imagem
            imagem = Image('registros/'+str(i+1)+'.jpg')
            imagem.width=710
            imagem.height=840

            #* Colocando as informações
            sheet[f'C{r}'] = self.titulos[i]
            sheet[f'A{k}'] = i+1
            sheet[f'J{k}'] = self.prioridades[i]
            sheet[f'B{t}'] = self.descricao[i]
            sheet[f'A{r}'] = "PROJETE 5D"
            sheet[f'B{k}'] = "TÍTULO"
            sheet[f'A{t}'] = "COMENTÁRIO"
            
            sheet.add_image(imagem,f'A{n}')

            #*Estilos das células
            sheet.merge_cells(f'C{r}:I{r}')
            sheet.merge_cells(f'A{r}:B{r}')
            sheet.merge_cells(f'B{t}:K{x}')
            sheet.merge_cells(f'B{k}:I{k}')
            sheet.merge_cells(f'J{r}:K{r}')
            sheet.merge_cells(f'J{k}:K{k}')
            sheet.merge_cells(f'A{t}:A{x}')

            k+=54
            r+=54
            t+=54
            n+=54
            x+=54

        lw.save(self.planilha)

Scrapper()       
