
#! Importações
from bs4 import BeautifulSoup as bs
from openpyxl.styles.borders import Side 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
import openpyxl as op
from openpyxl.drawing.image import Image
import openpyxl.styles as ops
import requests 

class Scrapper():
    def __init__(self):
        self.login = 'camaragibe@projete5d.com.br'
        self.password = '@timeprojete5d'
        self.longinurl = ('https://join.bimcollab.com/WebApp/Account/Login.aspx')
        self.planilha = 'registros/teste2.xlsx'
        Scrapper.coleta_de_dados(self)
        Scrapper.salvar_dados(self)
    def coleta_de_dados(self):
        #*Navegador
        conf = Options()
        conf.add_argument('--headless')
        navegador = webdriver.Chrome(options=conf)
        navegador.get(self.longinurl)

        #* Movimentação no sistema
        navegador.find_element_by_id('UserName').send_keys(self.login)
        navegador.find_element_by_id('Password').send_keys(self.password)
        navegador.find_element_by_css_selector('#LoginButton').click()
        sleep(2)
        selecionar_projeto = navegador.find_element_by_css_selector('#ProjectImage').click()
        selecionar_issue = navegador.find_element_by_css_selector('#HyperLinkIssue').click()
    
        #*Variáveis
        c=0
        self.descricao = []
        self.titulos = []
        self.prioridades = []
        imagens = []
        d = ''

        #*Coleta das informações
        a = navegador.find_elements_by_css_selector('.gridViewRow > .colTitle')
        navegador.find_element_by_css_selector('.gridViewRow > .colTitle').click()
    
        for i in range(len(a)):

            site = bs(navegador.page_source, 'html.parser')
        
            #* Coleta da descrição
            u = site.find_all('div',attrs={'class':'onerow commentContent innerPaddingComment'})
            for i in range(len(u)):
                desc = site.find_all('div',attrs={'class':'onerow commentContent innerPaddingComment'})[i]
                d = desc.p.text +'\n'+ d  
                
            self.descricao.append(d)    
            d = ''

            #* Coleta do título e prioridade
            titulo = site.find('span',attrs={'id':'LabelDescription'})
            prioridade = site.find('span', attrs={'id':'LabelPriority'})
            self.titulos.append(titulo.text)
            self.prioridades.append(prioridade.text)
            sleep(0.5)

            #* Coleta da imagem
            imagem = site.find('div',attrs={'class':'thumbnailImageFrame backgroundWhite'})
            a =imagem['style'] 
            imagens.append('https://join.bimcollab.com/' + a[21:-5])
            nome = 'registros/'+ str(i+1) +'.jpg'
            f = open(nome,'wb')
            response = requests.get('https://join.bimcollab.com/' + a[21:-5])
            f.write(response.content)
            f.close
            sleep(2)

            navegador.find_element_by_css_selector('#HyperLinkNext').click()

            c += 1
            if c == 5:
                break

        print(f"A quantidade de elementos é de {a}")
    def salvar_dados(self):
        lw = op.load_workbook(self.planilha)
        sheet = lw.active

        k = 1
        r = 2
        t = 3
        x = 8
        n = 10
        for i in range(len(self.titulos)):
       
            sheet.merge_cells(f'C{r}:I{r}')
            sheet.merge_cells(f'A{r}:B{r}')
            sheet.merge_cells(f'B{t}:K{x}')

            imagem = Image('registros/'+str(i+1)+'.jpg')
            imagem.width=700
            imagem.height=830
            
            borda = ops.Border(left=Side(border_style='thick',color='00000000'),
                                right=Side(border_style='thick',color='00000000'),
                                top=Side(border_style='thick',color='00000000'),
                                bottom=Side(border_style='thick',color='00000000'),
                                )

            fonte = ops.Font(name='Montserrat',
                            size=11)

            sheet['C' + str(r)] = self.titulos[i]
            sheet['A' + str(k)] = i+1
            sheet['A' + str(r)] = self.prioridades[i]
            sheet['B' + str(t)] = self.descricao[i]
            sheet.add_image(imagem,'A'+str(n))

            #*Estilos das células
            #?C2:G2
            sheet[f'C{r}'].border = borda
            sheet[f'D{r}'].border = borda
            sheet[f'E{r}'].border = borda
            sheet[f'F{r}'].border = borda
            sheet[f'G{r}'].border = borda
            
            sheet[f'A{k}'].border = borda
            #?A2:B2
            sheet[f'A{r}'].border = borda
            sheet[f'B{r}'].border = borda

            
            sheet[f'B{t}'].border = borda

            sheet[f'C{r}'].font = fonte
            sheet['A' + str(k)].font = fonte
            sheet[f'A{r}'].font = fonte
            sheet[f'B{r}'].font = fonte
            sheet[f'B{t}'].font = fonte

            k+=54
            r+=54
            t+=54
            n+=54
            x+=54
        
        print(f"A quantidade de títulos é {len(self.titulos)}")
        lw.save(self.planilha)

Scrapper()       
