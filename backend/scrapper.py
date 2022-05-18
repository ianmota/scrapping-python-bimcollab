from multiprocessing.reduction import send_handle
from types import NoneType
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
import time
import openpyxl as op
from openpyxl.drawing.image import Image
import openpyxl.styles as ops
import requests 

class Scrapper_collect():
    def __init__(self,login:str,senha:str,filtro:int,ordem:int,incompatibilidades:str,local_save:str):
        self.login = login
        self.password = senha
        self.filtro_geral = filtro #1 para abertos 0 para todos
        self.ordem_geral = ordem #0-id 1-area 2-titulo
        self.incompatibilidades = []
        self.local_save = local_save
    
    def __str__(self) -> str:
        return("Scrapper")
    
    def __repr__(self) -> str:
        return("Scrapper")
    
    def abrir_navegador(self)->webdriver:
        conf = Options()
        #conf.add_argument('--headless')
        self.navegador = webdriver.Chrome(options=conf)
        self.navegador.get("https://join.bimcollab.com/WebApp/Account/Login.aspx")
        time.sleep(1)
       
    def add_login(self)->NoneType:
            if(self.navegador.find_element_by_id("email")):
                self.navegador.find_element_by_id("email").send_keys(self.login)
                self.navegador.find_element_by_css_selector("#login").click()
                time.sleep(1)
                        
    def add_senha(self)->NoneType:
        if(self.navegador.find_element_by_id("password")):
            self.navegador.find_element_by_id("password").send_keys(self.password)
            self.navegador.find_element_by_css_selector("#loginWithPassword").click()
            time.sleep(2)
               
    def add_company(self)->NoneType:
        if(self.navegador.find_element_by_id("saveCompanyInfo")):
            self.navegador.find_element_by_id("saveCompanyInfo").click()
            time.sleep(2)
               
    def select_project(self)->NoneType:
        if(self.navegador.find_element_by_css_selector('#ProjectImage')):
            self.navegador.find_element_by_css_selector('#ProjectImage').click()   
            time.sleep(1)

    def go_to_issues(self)->NoneType:
        if(self.navegador.find_element_by_css_selector('#HyperLinkIssue')):
            self.navegador.find_element_by_css_selector('#HyperLinkIssue').click()
            time.sleep(4)
            
    def filtro(self)->NoneType:
        if(self.filtro_geral==1):
            if(self.navegador.find_element_by_id("LinkButtonAllOpen")):
                self.navegador.find_element_by_id("LinkButtonAllOpen").click()
                time.sleep(1)
        elif(self.filtro_geral==0):
            if(self.navegador.find_element_by_id("LinkButtonAll")):
                self.navegador.find_element_by_id("LinkButtonAll").click()
                time.sleep(1)
    
    def coletar_numero(self)->list:
        numero=[]
        if(self.site.find('div',attrs={'id':'LabelId'})):
            d = self.site.find('div',attrs={'id':'LabelId'}).text
            numero.append(d)
        return(numero)
        
    def coletar_descricao(self)->list:
        descricao = []

        if(self.site.find('span', attrs={"id":"LabelDescription"})):
            if(self.site.find('span',attrs={"id":"LabelDescription"}).text == '-'):
                self.navegador.back() 
            else:
                d = self.site.find("span",attrs={"id":"LabelDescription"}).text
                descricao.append(d)
        return(descricao)
                
    def coletar_titulo(self)->list:
        titulo = []

        if(self.site.find('span',attrs={'id':'LabelTopic'})):
            d = self.site.find('span',attrs={'id':'LabelTopic'}).text
            titulo.append(d)
        return(titulo)
    
    def coletar_prioridade(self)->list:
        prioridade = []

        if(self.site.find('span', attrs={'id':'LabelPriority'})):
            d = self.site.find('span', attrs={'id':'LabelPriority'})
            prioridade.append(d)
        return(prioridade)
    
    def coletar_responsavel(self)->list:
        responsavel = []

        if(self.site.find('span', attrs={'id':'LabelAssignTo'})):
            d = self.site.find('span', attrs={'id':'LabelAssignTo'}).text
            responsavel.append(d)
        return(responsavel)
    
    def coletar_tipo(self)->list:
        tipo = []

        if(self.site.find('span', attrs={'id':'LabelType'})):
            d = self.site.find('span', attrs={'id':'LabelType'}).text
            tipo.append(d)
        return(tipo)
    
    def coletar_area(self)->list:
        area = []

        if(self.site.find('span', attrs={'id':'LabelArea'})):
            d = self.site.find('span', attrs={'id':'LabelArea'}).text
            area.append(d)
        return(area)
    
    def coletar_data(self)->list:
        data = []

        if(self.site.find('span', attrs={'id':'LabelDeadline'})):
            d = self.site.find('span', attrs={'id':'LabelDeadline'}).text
            data.append(d)
        return(data)
            
    def coletar_etiqueta(self)->list:
        etiqueta = []

        if(self.site.find('span', attrs={'id':'LabelLabel'})):
            d = self.site.find('span', attrs={'id':'LabelLabel'}).text
            etiqueta.append(d)
        return(etiqueta)
    
    def coletar_milestone(self)->list:
        milestone = []

        if(self.site.find('span', attrs={'id':'LabelMilestone'})):
            d = self.site.find('span', attrs={'id':'LabelMilestone'}).text
            milestone.append(d)
        return(milestone)
    
    def coletar_status(self)->list:
        status = []

        if(self.site.find('span', attrs={'id':'LabelStatus'})):
            d = self.site.find('span', attrs={'id':'LabelStatus'}).text
            status.append(d)
        return(status)
    
    def coletar_empreendimento(self)->list:
        empreendimento = []

        if(self.site.find('span', attrs={'id':'LabelCurrentProject'})):
            d = self.site.find('span', attrs={'id':'LabelCurrentProject'}).text
            empreendimento.append(d)
        return(empreendimento)
    
    def coletar_comentarios_imagens(self)->list:
        comentario = []
        imagem = []
        elementos = self.site.find_all("div",attrs={"class":"onerow commentContent innerPaddingComment"})
        for i in range(len(elementos)):
            d = self.site.find_all("div",attrs={"class":"onerow commentContent innerPaddingComment"})[i]
            comentario.append(d)

            issueimg = self.site.find_all('a',attrs={"class":"viewpointIssueImgLink previewLeft"})
            for i in issueimg:

                issuelink = i["href"]
                imagem.append(issuelink[-12:])
                f = open(f"{self.local_save}/{issuelink[-12:]}.png","wb")
                f.write(requests.get(f"https://join.bimcollab.com{issuelink}").content)
                f.close() 
            return([comentario,imagem])
        
    def selecao_incompatibilidades(self):
        lista_issue = self.navegador.find_elements_by_css_selector(".gridViewRow > .colTitle")

        if self.incompatibilidades:
            
            filtro_incompatibilidades = []
            for i in self.incompatibilidades:
                a = ''
                for j in i:
                    if j == '-':
                        b = a
                        a = ''
                    else:
                        a = a + j
                
                if b:
                    for i in range(int(b)-1,int(a)):
                        filtro_incompatibilidades.append(i)
                    b = ''
                else:
                    filtro_incompatibilidades.append(int(a))
        
        if not self.incompatibilidades:
            filtro_incompatibilidades = range(len(lista_issue))
        return(filtro_incompatibilidades)
    
    def coletar_dados(self):

        for i in self.selecao_incompatibilidades():
            self.navegador.find_element_by_css_selector(f"#LabelIndex_{i}").click()

            self.site = bs(self.navegador.page_source, 'html.parser')
            
            dados = {
                "descricao": self.coletar_descricao(),
                "prioridade":self.coletar_prioridade(),
                "titulo": self.coletar_titulo(),
                "tipo": self.coletar_tipo(),
                "responsavel": self.coletar_responsavel(),
                "etiqueta": self.coletar_etiqueta(),
                "area":self.coletar_area(),
                "data": self.coletar_data(),
                "milestone": self.coletar_milestone(),
                "status": self.coletar_status(),
                "empreendimento": self.coletar_empreendimento(),
                "imagens": self.coletar_comentarios_imagens()
            }
            
            self.navegador.back()
            time.sleep(1)
            
        return(dados)
    
    def ordem(self):

        if(self.ordem_geral==0):
            site_issue = bs(self.navegador.page_source, "html.parser")
            verificador = site_issue.find("span",attrs={"id":"LabelIndex_0"})

            if(verificador.text != str(self.filtro_incompatibilidades[0]+1)):
                if(self.navegador.find_element_by_id("LabelHeaderNr")):
                    self.navegador.find_element_by_id("LabelHeaderNr").click()
                    self.navegador.refresh()
                    time.sleep(1)
            else: 
                self.VerificadorOrdem = True


        if(self.ordem_geral==1):
            if(self.navegador.find_element_by_id("LabelHeaderArea")):
                self.navegador.find_element_by_id("LabelHeaderArea").click()
                time.sleep(1)
        if(self.ordem_geral==2):
            if(self.navegador.find_element_by_id("LabelHeaderTitle")):
                self.navegador.find_element_by_id("LabelHeaderTitle").click()
                time.sleep(1)                                                                                                                     

    def salvar_dados(self):
        lw = op.Workbook()
        sheet = lw.active

        k = 1
        r = 2
        t = 3
        x = 8
        n = 10
        count = 0
        for i in range(len(self.titulos)):

            thick = ops.Side(border_style='thin',color='00000000')
            fonte = ops.Font(name='Montserrat',size=11)
            alinhamento = ops.Alignment(horizontal="distributed",vertical="distributed")
            alinhamento3 = ops.Alignment(horizontal="left",vertical="distributed")
            alinhamento2 = ops.Alignment(horizontal="distributed",vertical="distributed",text_rotation=90)

            #?C2:I2
            sheet[f'C{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)            
            sheet[f'C{r}'].font = fonte
            sheet[f'C{r}'].Alignment = alinhamento3
            #? A1
            sheet[f'A{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{k}'].font = fonte
            sheet[f'A{k}'].Alignment = alinhamento
            #? A2:B2
            sheet[f'A{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{r}'].font = fonte
            sheet[f'A{r}'].Alignment = alinhamento3
            #? B3:K8
            sheet[f'B{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{t}'].font = fonte
            sheet[f'B{t}'].Alignment = alinhamento3

            sheet[f'B{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{k}'].font = fonte
            sheet[f'B{k}'].Alignment = alinhamento3

            sheet[f'H{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{k}'].font = fonte
            sheet[f'H{k}'].Alignment = alinhamento

            sheet[f'H{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{r}'].font = fonte
            sheet[f'H{r}'].Alignment = alinhamento

            sheet[f'A{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{t}'].font = fonte
            sheet[f'A{t}'].Alignment = alinhamento2

            #* Estilo de imagem
            imagem = Image('data/'+str(i+1)+'.jpg')
            #imagem.width=599
            #imagem.height=763

            #* Colocando as informações
            sheet[f'C{r}'] = self.titulos[i]
            sheet[f'A{k}'] = count+1
            sheet[f'H{k}'] = self.prioridades[i]
            sheet[f'B{t}'] = self.descricao[i]
            sheet[f'A{r}'] = "PROJETE 5D"
            sheet[f'B{k}'] = self.nomes[i]
            sheet[f'A{t}'] = "COMENTÁRIO"
            
            sheet.add_image(imagem,f'A{n}')

            #*Estilos das células
            sheet.merge_cells(f'C{r}:G{r}')
            sheet.merge_cells(f'A{r}:B{r}')
            sheet.merge_cells(f'B{t}:I{x}')
            sheet.merge_cells(f'B{k}:G{k}')
            sheet.merge_cells(f'H{r}:I{r}')
            sheet.merge_cells(f'H{k}:I{k}')
            sheet.merge_cells(f'A{t}:A{x}')

            k+=50
            r+=50
            t+=50
            n+=50
            x+=50
            count+=1
        
        nome_planilha = "data/"
        for i in self.login:
            if i == "@":
                break
            nome_planilha = nome_planilha + i 
        
        nome_planilha = nome_planilha + ".xlsx"
        lw.save(nome_planilha)    
                
