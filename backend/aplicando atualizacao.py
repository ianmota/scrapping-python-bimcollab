from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
import time
import openpyxl as op
from openpyxl.drawing.image import Image
import openpyxl.styles as ops
import requests 

class Scrapper():
    def __init__(self):
        self.login = "oscarfreire@projete5d.com.br"
        self.password = "@Timeprojete5d"
        self.filtro_geral = 1 #1 para abertos 0 para todos
        self.ordem_geral = 0 #0 id 1-area 2-titulo
        self.incompatibilidades = ["35-38"]
        self.local_save = "data"
        self.start_relatorio()
    def abrir_navegador(self):

        conf = Options()
        #conf.add_argument('--headless')
        self.navegador = webdriver.Chrome(options=conf)
        self.navegador.get("https://join.bimcollab.com/WebApp/Account/Login.aspx")
    def add_login(self):
        try:
            if(self.navegador.find_element_by_id("email")):
                self.navegador.find_element_by_id("email").send_keys(self.login)
                self.navegador.find_element_by_css_selector("#login").click()
                time.sleep(1)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            time.sleep(1)
            print("Erro no login")
            self.VerificadorEmail = True        
    def add_senha(self):
        try:
            if(self.navegador.find_element_by_id("password")):
                self.navegador.find_element_by_id("password").send_keys(self.password)
                self.navegador.find_element_by_css_selector("#loginWithPassword").click()
                time.sleep(2)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            time.sleep(1)
            print("Erro na senha")
            self.VerificadorSenha = True    
    def add_company(self):
        try:
            if(self.navegador.find_element_by_id("saveCompanyInfo")):
                self.navegador.find_element_by_id("saveCompanyInfo").click()
                time.sleep(2)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            print("Erro na definição da companhia")
            time.sleep(1)
            self.VerificadorCompany = True
    def select_project(self):
        try:
            if(self.navegador.find_element_by_css_selector('#ProjectImage')):
                self.navegador.find_element_by_css_selector('#ProjectImage').click()   
                time.sleep(1)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            print("Erro na seleção do projeto")
            time.sleep(1)
            self.VerificadorProjeto = True
    def go_to_issues(self):
        try:
            if(self.navegador.find_element_by_css_selector('#HyperLinkIssue')):
                self.navegador.find_element_by_css_selector('#HyperLinkIssue').click()
                time.sleep(4)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            print("Erro na seleção do projeto")
            time.sleep(1)
            self.VerificadorIssues = True
    def filtro(self):
        try:
            if(self.incompatibilidades):
                self.filtro_geral = 0

            if(self.filtro_geral==1):
                if(self.navegador.find_element_by_id("LinkButtonAllOpen")):
                    self.navegador.find_element_by_id("LinkButtonAllOpen").click()
                    time.sleep(1)
            elif(self.filtro_geral==0):
                if(self.navegador.find_element_by_id("LinkButtonAll")):
                    self.navegador.find_element_by_id("LinkButtonAll").click()
                    time.sleep(1)
        except selenium.common.exceptions.NoSuchElementException:
            self.navegador.refresh()
            print("Erro na seleção do projeto")
            time.sleep(1)
            self.VerificadorFiltro = True
    def ordem(self):
        if(self.ordem_geral==0):
            if(self.navegador.find_element_by_id("LabelHeaderNr")):
                self.navegador.find_element_by_id("LabelHeaderNr").click()
                time.sleep(1)
        if(self.ordem_geral==1):
            if(self.navegador.find_element_by_id("LabelHeaderArea")):
                self.navegador.find_element_by_id("LabelHeaderArea").click()
                time.sleep(1)
        if(self.ordem_geral==2):
            if(self.navegador.find_element_by_id("LabelHeaderTitle")):
                self.navegador.find_element_by_id("LabelHeaderTitle").click()
                time.sleep(1)
    def coletar_descricao(self):
        self.descricao = []

        if(self.site.find('span', attrs={"id":"LabelDescription"})):
            if(self.site.find('span',attrs={"id":"LabelDescription"}).text == '-'):
                self.navegador.back() 
            else:
                d = self.site.find("span",attrs={"id":"LabelDescription"}).text
                self.descricao.append(d)
    def coletar_titulo(self):
        self.titulo = []

        if(self.site.find('span',attrs={'id':'LabelTopic'})):
            d = self.site.find('span',attrs={'id':'LabelTopic'}).text
            self.titulo.append(d)
    def coletar_prioridade(self):
        self.prioridade = []

        if(self.site.find('span', attrs={'id':'LabelPriority'})):
            d = self.site.find('span', attrs={'id':'LabelPriority'})
            self.prioridade.append(d)
    def coletar_responsavel(self):
        self.responsavel = []

        if(self.site.find('span', attrs={'id':'LabelAssignTo'})):
            d = self.site.find('span', attrs={'id':'LabelAssignTo'}).text
            self.responsavel.append(d)
    def coletar_tipo(self):
        self.tipo = []

        if(self.site.find('span', attrs={'id':'LabelType'})):
            d = self.site.find('span', attrs={'id':'LabelType'}).text
            self.tipo.append(d)
    def coletar_area(self):
        self.area = []

        if(self.site.find('span', attrs={'id':'LabelArea'})):
            d = self.site.find('span', attrs={'id':'LabelArea'}).text
            self.area.append(d)
    def coletar_data(self):
        self.data = []

        if(self.site.find('span', attrs={'id':'LabelDeadline'})):
            d = self.site.find('span', attrs={'id':'LabelDeadline'}).text
            self.data.append(d)
    def coletar_etiqueta(self):
        self.etiqueta = []

        if(self.site.find('span', attrs={'id':'LabelLabel'})):
            d = self.site.find('span', attrs={'id':'LabelLabel'}).text
            self.etiqueta.append(d)
    def coletar_milestone(self):
        self.milestone = []

        if(self.site.find('span', attrs={'id':'LabelMilestone'})):
            d = self.site.find('span', attrs={'id':'LabelMilestone'}).text
            self.milestone.append(d)
    def coletar_status(self):
        self.status = []

        if(self.site.find('span', attrs={'id':'LabelStatus'})):
            d = self.site.find('span', attrs={'id':'LabelStatus'}).text
            self.status.append(d)
    def coletar_empreendimento(self):
        self.empreendimento = []

        if(self.site.find('span', attrs={'id':'LabelCurrentProject'})):
            d = self.site.find('span', attrs={'id':'LabelCurrentProject'}).text
            self.empreendimento.append(d)
    def coletar_comentarios_imagens(self):
        self.comentario = []
        self.imagem = []
        elementos = self.site.find_all("div",attrs={"class":"onerow commentContent innerPaddingComment"})
        for i in range(len(elementos)):
            d = self.site.find_all("div",attrs={"class":"onerow commentContent innerPaddingComment"})[i]
            self.comentario.append(d)

            issueimg = self.site.find_all('a',attrs={"class":"viewpointIssueImgLink previewLeft"})
            for i in issueimg:

                issuelink = i["href"]
                self.imagem.append(issuelink)
                f = open(f"{self.local_save}{issuelink}.png","wb")
                f.write(requests.get(f"https://join.bimcollab.com{issuelink}").content)
                f.close()
    def selecao_incompatibilidades(self):
        lista_issue = self.navegador.find_elements_by_css_selector(".gridViewRow > .colTitle")

        if self.incompatibilidades:
            
            self.filtro_incompatibilidades = []
            for i in self.incompatibilidades:
                a = ''
                for j in i:
                    if j == '-':
                        b = a
                        a = ''
                    else:
                        a = a + j
                
                if b:
                    for i in range(int(b)-1,int(a)):
                        self.filtro_incompatibilidades.append(i)
                    b = ''
                else:
                    self.filtro_incompatibilidades.append(int(a))

        if not self.incompatibilidades:
            self.filtro = range(len(lista_issue))
    def coletar_dados(self):
        self.selecao_incompatibilidades()

        for i in self.filtro_incompatibilidades:
            self.navegador.find_element_by_css_selector(f"#LabelIndex_{i}").click()
            self.site = bs(self.navegador.page_source, 'html.parser')

            self.coletar_descricao()
            self.coletar_prioridade()
            self.coletar_titulo()
            self.coletar_tipo()
            self.coletar_responsavel()
            self.coletar_etiqueta()
            self.coletar_area()
            self.coletar_data()
            self.coletar_milestone()
            self.coletar_status()
            self.coletar_empreendimento()
            self.coletar_comentarios_imagens()
            self.navegador.back()
            time.sleep(1)
   
    def start_relatorio(self):
        self.VerificadorEmail = True
        self.VerificadorSenha = False
        self.VerificadorCompany = False
        self.VerificadorProjeto = False
        self.VerificadorIssues = False
        self.VerificadorFiltro = False   
        VerificadorOrdem = False
        VerificadorColeta = False
        self.abrir_navegador()

        time.sleep(1)
        while True:
            if(self.VerificadorEmail):
                self.add_login()
                self.VerificadorEmail = False
                self.VerificadorSenha = True
            if(self.VerificadorSenha):
                self.add_senha()
                self.VerificadorSenha = False
                self.VerificadorCompany = True
            if(self.VerificadorCompany):
                self.add_company()
                self.VerificadorCompany = False   
                self.VerificadorProjeto = True                 
            if(self.VerificadorProjeto):
                self.select_project()
                self.VerificadorProjeto = False
                self.VerificadorIssues = True
            if(self.VerificadorIssues):
                self.go_to_issues()
                self.VerificadorIssues = False
                self.VerificadorFiltro = True
            if(self.VerificadorFiltro):
                self.filtro()
                self.VerificadorFiltro = False
                VerificadorOrdem = True
            if(VerificadorOrdem):
                self.ordem()
                VerificadorOrdem = False
                VerificadorColeta = True
            if(VerificadorColeta):
                self.coletar_dados()
                VerificadorColeta = False
    

    def salvar_dados(self):
        lw = op.Workbook()
        sheet = lw.active

        k = 1
        r = 2
        t = 3
        x = 8
        n = 10
        count = 0
        for i in range(len(self.titulos)):

            thick = ops.Side(border_style='thin',color='00000000')
            fonte = ops.Font(name='Montserrat',size=11)
            alinhamento = ops.Alignment(horizontal="distributed",vertical="distributed")
            alinhamento3 = ops.Alignment(horizontal="left",vertical="distributed")
            alinhamento2 = ops.Alignment(horizontal="distributed",vertical="distributed",text_rotation=90)

            #?C2:I2
            sheet[f'C{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)            
            sheet[f'C{r}'].font = fonte
            sheet[f'C{r}'].Alignment = alinhamento3
            #? A1
            sheet[f'A{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{k}'].font = fonte
            sheet[f'A{k}'].Alignment = alinhamento
            #? A2:B2
            sheet[f'A{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{r}'].font = fonte
            sheet[f'A{r}'].Alignment = alinhamento3
            #? B3:K8
            sheet[f'B{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{t}'].font = fonte
            sheet[f'B{t}'].Alignment = alinhamento3

            sheet[f'B{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{k}'].font = fonte
            sheet[f'B{k}'].Alignment = alinhamento3

            sheet[f'H{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{k}'].font = fonte
            sheet[f'H{k}'].Alignment = alinhamento

            sheet[f'H{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{r}'].font = fonte
            sheet[f'H{r}'].Alignment = alinhamento

            sheet[f'A{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{t}'].font = fonte
            sheet[f'A{t}'].Alignment = alinhamento2

            #* Estilo de imagem
            imagem = Image('data/'+str(i+1)+'.jpg')
            #imagem.width=599
            #imagem.height=763

            #* Colocando as informações
            sheet[f'C{r}'] = self.titulos[i]
            sheet[f'A{k}'] = count+1
            sheet[f'H{k}'] = self.prioridades[i]
            sheet[f'B{t}'] = self.descricao[i]
            sheet[f'A{r}'] = "PROJETE 5D"
            sheet[f'B{k}'] = self.nomes[i]
            sheet[f'A{t}'] = "COMENTÁRIO"
            
            sheet.add_image(imagem,f'A{n}')

            #*Estilos das células
            sheet.merge_cells(f'C{r}:G{r}')
            sheet.merge_cells(f'A{r}:B{r}')
            sheet.merge_cells(f'B{t}:I{x}')
            sheet.merge_cells(f'B{k}:G{k}')
            sheet.merge_cells(f'H{r}:I{r}')
            sheet.merge_cells(f'H{k}:I{k}')
            sheet.merge_cells(f'A{t}:A{x}')

            k+=50
            r+=50
            t+=50
            n+=50
            x+=50
            count+=1
        
        nome_planilha = "data/"
        for i in self.login:
            if i == "@":
                break
            nome_planilha = nome_planilha + i 
        
        nome_planilha = nome_planilha + ".xlsx"
        lw.save(nome_planilha)    
                
Scrapper()