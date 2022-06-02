import openpyxl as op
from openpyxl.drawing.image import Image
import openpyxl.styles as ops
import pandas as pd

class DataSave():
    def __init__(self,database) -> None:
        """
        Salva todos os dados em um arquivo
        Args:
            database (ScrapperColect): dicionário com os dados
        """
        self.db = database 
    def salvar_dados_relatório(self):
        """
        Não usar, ainda estão em desenvolvimento
        """
        lw = op.Workbook()
        sheet = lw.active

        k = 1
        r = 2
        t = 3
        x = 8
        n = 10
        count = 0
        for i in range(len(self.titulos)):

            thick = ops.Side(border_style='thin',color='00000000')
            fonte = ops.Font(name='Montserrat',size=11)
            alinhamento = ops.Alignment(horizontal="distributed",vertical="distributed")
            alinhamento3 = ops.Alignment(horizontal="left",vertical="distributed")
            alinhamento2 = ops.Alignment(horizontal="distributed",vertical="distributed",text_rotation=90)

            #?C2:I2
            sheet[f'C{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)            
            sheet[f'C{r}'].font = fonte
            sheet[f'C{r}'].Alignment = alinhamento3
            #? A1
            sheet[f'A{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{k}'].font = fonte
            sheet[f'A{k}'].Alignment = alinhamento
            #? A2:B2
            sheet[f'A{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{r}'].font = fonte
            sheet[f'A{r}'].Alignment = alinhamento3
            #? B3:K8
            sheet[f'B{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{t}'].font = fonte
            sheet[f'B{t}'].Alignment = alinhamento3

            sheet[f'B{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'B{k}'].font = fonte
            sheet[f'B{k}'].Alignment = alinhamento3

            sheet[f'H{k}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{k}'].font = fonte
            sheet[f'H{k}'].Alignment = alinhamento

            sheet[f'H{r}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'H{r}'].font = fonte
            sheet[f'H{r}'].Alignment = alinhamento

            sheet[f'A{t}'].Border = ops.Border(left=thick,right=thick,top=thick,bottom=thick)
            sheet[f'A{t}'].font = fonte
            sheet[f'A{t}'].Alignment = alinhamento2

            #* Estilo de imagem
            imagem = Image('data/'+str(i+1)+'.jpg')
            #imagem.width=599
            #imagem.height=763

            #* Colocando as informações
            sheet[f'C{r}'] = self.titulos[i]
            sheet[f'A{k}'] = count+1
            sheet[f'H{k}'] = self.prioridades[i]
            sheet[f'B{t}'] = self.descricao[i]
            sheet[f'A{r}'] = "PROJETE 5D"
            sheet[f'B{k}'] = self.nomes[i]
            sheet[f'A{t}'] = "COMENTÁRIO"
            
            sheet.add_image(imagem,f'A{n}')

            #*Estilos das células
            sheet.merge_cells(f'C{r}:G{r}')
            sheet.merge_cells(f'A{r}:B{r}')
            sheet.merge_cells(f'B{t}:I{x}')
            sheet.merge_cells(f'B{k}:G{k}')
            sheet.merge_cells(f'H{r}:I{r}')
            sheet.merge_cells(f'H{k}:I{k}')
            sheet.merge_cells(f'A{t}:A{x}')

            k+=50
            r+=50
            t+=50
            n+=50
            x+=50
            count+=1
        
        nome_planilha = "data/"
        for i in self.login:
            if i == "@":
                break
            nome_planilha = nome_planilha + i 
        
        nome_planilha = nome_planilha + ".xlsx"
        lw.save(nome_planilha)    
    
def dbBuildCSV(dic:dict,localSave:str):
    index = dic.pop("ID")
    
    database = pd.DataFrame(dic,index=index)
    database.to_csv(path_or_buf=localSave,index= False,sep=";")
    return(database)